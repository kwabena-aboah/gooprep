"""Excel and PDF export service for admin reports."""
import io
from datetime import datetime


# ── Excel ─────────────────────────────────────────────────────────
def _apply_header_style(ws, row_num, headers):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    RED   = 'E63900'
    WHITE = 'FFFFFF'
    header_font = Font(bold=True, color=WHITE, size=11)
    header_fill = PatternFill('solid', fgColor=RED)
    center      = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center
        cell.border     = thin_border


def _auto_width(ws, headers):
    from openpyxl.utils import get_column_letter
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(h)) + 4)


def export_users_excel(queryset) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Users'
    ws.row_dimensions[1].height = 22
    headers = ['ID','Full Name','Email','Role','Plan','City','Points','Level','Streak','Active','Joined']
    _apply_header_style(ws, 1, headers)
    for row, u in enumerate(queryset, 2):
        ws.cell(row=row, column=1,  value=u.id)
        ws.cell(row=row, column=2,  value=u.get_full_name())
        ws.cell(row=row, column=3,  value=u.email)
        ws.cell(row=row, column=4,  value=u.role)
        ws.cell(row=row, column=5,  value=u.subscription_plan)
        ws.cell(row=row, column=6,  value=u.city)
        ws.cell(row=row, column=7,  value=u.total_points)
        ws.cell(row=row, column=8,  value=u.level)
        ws.cell(row=row, column=9,  value=u.streak_days)
        ws.cell(row=row, column=10, value='Yes' if u.is_active else 'No')
        ws.cell(row=row, column=11, value=u.date_joined.strftime('%Y-%m-%d') if u.date_joined else '')
        if row % 2 == 0:
            from openpyxl.styles import PatternFill
            alt = PatternFill('solid', fgColor='FFF1ED')
            for c in range(1, len(headers)+1): ws.cell(row=row, column=c).fill = alt
    _auto_width(ws, headers)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def export_lessons_excel(queryset) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Lessons'
    headers = ['ID','Tutor','Student','Subject','Type','Status','Start Time','Duration(min)','Price(GHS)','Payment','On Behalf','Booker']
    _apply_header_style(ws, 1, headers)
    for row, l in enumerate(queryset.select_related('tutor','student','subject'), 2):
        ws.cell(row=row, column=1,  value=l.id)
        ws.cell(row=row, column=2,  value=l.tutor.get_full_name())
        ws.cell(row=row, column=3,  value=l.student.get_full_name())
        ws.cell(row=row, column=4,  value=l.subject.name if l.subject else '')
        ws.cell(row=row, column=5,  value=l.lesson_type)
        ws.cell(row=row, column=6,  value=l.status)
        ws.cell(row=row, column=7,  value=l.start_time.strftime('%Y-%m-%d %H:%M') if l.start_time else '')
        ws.cell(row=row, column=8,  value=l.duration_minutes)
        ws.cell(row=row, column=9,  value=float(l.price))
        ws.cell(row=row, column=10, value=l.payment_status)
        ws.cell(row=row, column=11, value='Yes' if l.booked_on_behalf else 'No')
        ws.cell(row=row, column=12, value=f'{l.booker_name} ({l.booker_relationship})' if l.booked_on_behalf else '')
    _auto_width(ws, headers)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def export_revenue_excel(queryset) -> bytes:
    import openpyxl
    from decimal import Decimal
    from django.conf import settings
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Revenue'
    headers = ['ID','Payer','Amount(GHS)','Platform Fee','Tutor Net','Method','Status','Reference','Date']
    _apply_header_style(ws, 1, headers)
    total_gross = total_fee = Decimal(0)
    for row, t in enumerate(queryset.select_related('payer'), 2):
        fee = t.amount * Decimal(str(settings.PLATFORM_COMMISSION))
        net = t.amount * (Decimal('1') - Decimal(str(settings.PLATFORM_COMMISSION)) )
        total_gross += t.amount; total_fee += fee
        ws.cell(row=row, column=1, value=t.id)
        ws.cell(row=row, column=2, value=t.payer.get_full_name())
        ws.cell(row=row, column=3, value=float(t.amount))
        ws.cell(row=row, column=4, value=float(fee))
        ws.cell(row=row, column=5, value=float(net))
        ws.cell(row=row, column=6, value=t.payment_method)
        ws.cell(row=row, column=7, value=t.status)
        ws.cell(row=row, column=8, value=t.paystack_ref or '')
        ws.cell(row=row, column=9, value=t.created_at.strftime('%Y-%m-%d') if t.created_at else '')
    total_row = queryset.count() + 2
    from openpyxl.styles import Font
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=float(total_gross)).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=float(total_fee)).font   = Font(bold=True)
    _auto_width(ws, headers)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def export_tutors_excel(queryset) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Tutors'
    headers = ['ID','Name','Email','City','Rate(GHS)','Avg Rating','Reviews','Lessons','Students','Status','Featured','Joined']
    _apply_header_style(ws, 1, headers)
    for row, tp in enumerate(queryset.select_related('user'), 2):
        ws.cell(row=row, column=1,  value=tp.id)
        ws.cell(row=row, column=2,  value=tp.user.get_full_name())
        ws.cell(row=row, column=3,  value=tp.user.email)
        ws.cell(row=row, column=4,  value=tp.user.city)
        ws.cell(row=row, column=5,  value=float(tp.hourly_rate))
        ws.cell(row=row, column=6,  value=float(tp.average_rating))
        ws.cell(row=row, column=7,  value=tp.total_reviews)
        ws.cell(row=row, column=8,  value=tp.total_lessons)
        ws.cell(row=row, column=9,  value=tp.total_students)
        ws.cell(row=row, column=10, value=tp.approval_status)
        ws.cell(row=row, column=11, value='Yes' if tp.is_featured else 'No')
        ws.cell(row=row, column=12, value=tp.created_at.strftime('%Y-%m-%d') if tp.created_at else '')
    _auto_width(ws, headers)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def export_referrals_excel(queryset) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Referrals'
    headers = ['ID','Full Name','Email','Role','Referrer Name','Referrer Notes','Joined']
    _apply_header_style(ws, 1, headers)
    for row, u in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=u.id)
        ws.cell(row=row, column=2, value=u.get_full_name())
        ws.cell(row=row, column=3, value=u.email)
        ws.cell(row=row, column=4, value=u.role)
        ws.cell(row=row, column=5, value=u.referrer_name)
        ws.cell(row=row, column=6, value=u.referrer_notes)
        ws.cell(row=row, column=7, value=u.date_joined.strftime('%Y-%m-%d') if u.date_joined else '')
    _auto_width(ws, headers)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ── PDF ────────────────────────────────────────────────────────────
def _pdf_base(title: str, headers: list, rows: list) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf  = io.BytesIO()
    page = landscape(A4) if len(headers) > 6 else A4
    doc  = SimpleDocTemplate(buf, pagesize=page,
                              topMargin=2.5*cm, bottomMargin=1.5*cm,
                              leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    hdr_s  = ParagraphStyle('hdr',  parent=styles['Normal'], fontSize=9,  textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_CENTER)
    cell_s = ParagraphStyle('cell', parent=styles['Normal'], fontSize=8,  leading=11)
    table_data = [[Paragraph(h, hdr_s) for h in headers]]
    for r in rows:
        table_data.append([Paragraph(str(c) if c is not None else '', cell_s) for c in r])
    col_w = (page[0] - 2*cm) / len(headers)
    tbl   = Table(table_data, colWidths=[col_w]*len(headers), repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#E63900')),
        ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
        ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0),  9),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),  [colors.white, colors.HexColor('#FFF1ED')]),
        ('FONTNAME',      (0,1),(-1,-1),  'Helvetica'),
        ('FONTSIZE',      (0,1),(-1,-1),  8),
        ('GRID',          (0,0),(-1,-1),  0.4, colors.HexColor('#E2E8F0')),
        ('VALIGN',        (0,0),(-1,-1),  'TOP'),
        ('TOPPADDING',    (0,0),(-1,-1),  3),
        ('BOTTOMPADDING', (0,0),(-1,-1),  3),
    ]))

    def on_page(canvas, doc, title=title):
        canvas.saveState()
        canvas.setFillColor(colors.HexColor('#E63900'))
        canvas.rect(0, page[1]-1.8*cm, page[0], 1.8*cm, fill=True, stroke=False)
        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica-Bold', 14)
        canvas.drawString(1*cm, page[1]-1.2*cm, f'GOOPREP — {title.upper()}')
        canvas.setFont('Helvetica', 8)
        canvas.drawRightString(page[0]-1*cm, page[1]-1.2*cm,
                               f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}')
        canvas.setFillColor(colors.HexColor('#64748B'))
        canvas.setFont('Helvetica', 7)
        canvas.drawString(1*cm, 0.6*cm, 'Gooprep by Sikaba Systems · Ghana 🇬🇭')
        canvas.drawRightString(page[0]-1*cm, 0.6*cm, f'Page {doc.page}')
        canvas.restoreState()

    doc.build([tbl], onFirstPage=on_page, onLaterPages=on_page)
    return buf.getvalue()


def export_users_pdf(qs) -> bytes:
    headers = ['ID','Name','Email','Role','Plan','City','Active','Joined']
    rows = [[u.id,u.get_full_name(),u.email,u.role,u.subscription_plan,u.city,
             'Yes' if u.is_active else 'No',
             u.date_joined.strftime('%Y-%m-%d') if u.date_joined else ''] for u in qs]
    return _pdf_base('User Report', headers, rows)


def export_lessons_pdf(qs) -> bytes:
    headers = ['ID','Tutor','Student','Subject','Status','Start','Price(GHS)','On Behalf']
    rows = [[l.id,l.tutor.get_full_name(),l.student.get_full_name(),
             l.subject.name if l.subject else '',l.status,
             l.start_time.strftime('%d %b %Y %H:%M') if l.start_time else '',
             f'GHS {l.price}','Yes' if l.booked_on_behalf else 'No']
            for l in qs.select_related('tutor','student','subject')]
    return _pdf_base('Lesson Report', headers, rows)


def export_revenue_pdf(qs) -> bytes:
    from django.conf import settings
    commission = float(settings.PLATFORM_COMMISSION)
    headers = ['ID','Payer','Amount','Platform Fee','Tutor Net','Method','Status','Date']
    rows = [[t.id,t.payer.get_full_name(),f'GHS {t.amount}',
             f'GHS {float(t.amount)*commission:.2f}',f'GHS {float(t.amount)*(1-commission):.2f}',
             t.payment_method,t.status,
             t.created_at.strftime('%d %b %Y') if t.created_at else '']
            for t in qs.select_related('payer')]
    return _pdf_base('Revenue Report', headers, rows)


def export_tutors_pdf(qs) -> bytes:
    headers = ['ID','Name','Email','Rate(GHS)','Rating','Lessons','Status','Joined']
    rows = [[tp.id,tp.user.get_full_name(),tp.user.email,f'GHS {tp.hourly_rate}',
             str(tp.average_rating),tp.total_lessons,tp.approval_status,
             tp.created_at.strftime('%Y-%m-%d') if tp.created_at else '']
            for tp in qs.select_related('user')]
    return _pdf_base('Tutor Report', headers, rows)